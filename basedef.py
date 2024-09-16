from typing import List
import openpyxl
from difflib import SequenceMatcher
from datetime import datetime, timedelta

days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
lessons = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]

class LessonKP11():
    teacher = ""
    cabinet = ""
    lesson_num = ""
    group = ""
    lesson_name = ""
    lesson_day = ""
    coordinate = ""

    def __str__(self) -> str:
        return f"{self.teacher} {self.cabinet} {self.lesson_num} {self.group} {self.lesson_name} {self.lesson_day}"
    
    def __eq__(self, other):
        if isinstance(other, LessonKP11):
            tech_true = self.teacher == other.teacher and self.lesson_num == other.lesson_num and self.lesson_day == other.lesson_day
            cabinet_true = self.cabinet == other.cabinet and self.lesson_num == other.lesson_num and self.lesson_day== other.lesson_day
            
            return (tech_true or cabinet_true) and (self.teacher != "Куратор группы")
        return False
    
    def checkNakTeacher(self, other):
        if isinstance(other, LessonKP11):
            tech_true = self.teacher == other.teacher and self.lesson_num == other.lesson_num and self.lesson_day == other.lesson_day
            return tech_true  and self.teacher != "Куратор группы"
        return False
    
    def checkNakCabinet(self, other):
        if isinstance(other, LessonKP11):
            cabinet_true = self.cabinet == other.cabinet and self.lesson_num == other.lesson_num and self.lesson_day == other.lesson_day
            return cabinet_true and self.teacher != "Куратор группы"
        return False

    def check_teacher(self, out_teacher):
        matcher = SequenceMatcher(None, out_teacher, self.teacher)
        return True if matcher.ratio() > 0.83 else False

def check_teachers(teacher1, teacher2):
    matcher = SequenceMatcher(None, teacher1, teacher2)
    return True if matcher.ratio() > 0.83 else False

def get_group(ws, row_id, col_id):
    for row in ws.iter_rows(min_row=row_id, max_row=row_id):
        for cell in row:
            if cell.value is not None and cell.column == col_id:
                return cell.value
            
def get_groups(ws, row_id = 4):
    groups = []
    for row in ws.iter_rows(min_row=row_id, max_row=row_id):
        for cell in row:
            if cell.value is not None:
                groups.append(cell.value)
    return groups

def get_day(ws, col_id, row_id):
    new_days = []
    for col in ws.iter_cols(min_col=col_id, max_col=col_id, max_row=row_id):
        for cell in col:
            if cell.value is not None and cell.value in days:
                #print(cell.value)
                new_days.append(cell.value)
    return new_days[-1] if new_days is not None else "понедельник"

def get_lesson_num(ws, col_id, row_id) -> int:
    for col in ws.iter_cols(min_col=col_id, max_col=col_id, max_row=row_id, min_row=row_id-1):
        for cell in col:
            if cell.value is not None and cell.value in range(1, 11):
                return int(cell.value)

def get_teachers(ws):
    teachers = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "/" in str(cell.value):
                #print(len(teachers))
                if len(teachers) == 0:
                    teachers.append(cell.value.split("/")[-1].strip())
                elif len(teachers) == 1:
                    if not check_teachers(teachers[0], cell.value.split("/")[-1].strip()):
                        teachers.append(cell.value.split("/")[-1].strip())
                else:
                    c = 0
                    for i in teachers:
                        if check_teachers(i, cell.value.split("/")[-1].strip()):
                            c+=1
                        
                    if c > 0:
                        continue
                    else:
                        teachers.append(cell.value.split("/")[-1].strip())
                
    return set(teachers)

def get_cabinets(ws):
    cabinets = []
    cab_errors = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "/" in str(cell.value) or str(cell.value) == "Разговоры о важном":
                cabinets.append(cell.offset(0, 1).value)
                if len(str(cell.offset(0, 1).value)) < 3 or str(cell.offset(0, 1).value) == None:
                    cab_errors.append(cell.offset(0, 1).coordinate)
    return [set(cabinets), cab_errors]

def get_unic_lesson(ws):
    lessons_arr = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "/" in str(cell.value) or str(cell.value) == "Разговоры о важном":
                if "(1 п" in cell.value or "(2 п" in cell.value:
                    str(cell.value).replace("(1 п/гр)", "").replace("(2 п/гр)", "")
                lessons_arr.append(cell.value.split("/")[0] if "/" in str(cell.value) else cell.value)
    return set(lessons_arr)

def get_lesson(ws) -> List[LessonKP11]:
    lessons_arr = []
    teachers = get_teachers(ws)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "/" in str(cell.value) or str(cell.value) == "Разговоры о важном":
                if cell.offset(0, 1).value is not None:
                    lesson = LessonKP11()
                    lesson.cabinet = cell.offset(0, 1).value
                    lesson.group = get_group(ws, 4, cell.column)
                    lesson.lesson_day = get_day(ws, 1, cell.row)
                    teach = "Куратор группы"
                    if "/" in str(cell.value).strip():
                        teach = str(cell.value).split("/")[-1].strip() 
                        for t in teachers:
                            if check_teachers(teach, t):
                                teach = t
                                break
                    lesson.teacher = teach
                    lesson.lesson_num = get_lesson_num(ws, 2, cell.row)
                    lesson.lesson_name = cell.value.split("/")[0] if "/" in str(cell.value) else cell.value
                    lesson.coordinate = cell.coordinate
                    if "(1 п" in str(lesson.lesson_name) or "(2 п" in str(lesson.lesson_name):
                        lesson.lesson_name += "/гр)"
                    lessons_arr.append(lesson)
    return lessons_arr

