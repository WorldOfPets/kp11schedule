import openpyxl
import sys
from dbfunc import find_less, find_cabinet, find_teacher
from openpyxl.styles import PatternFill

def check_teachers(filePath):
    wb = openpyxl.load_workbook(filePath, read_only=False, keep_vba=True)

    ws = wb.active
    try:
        for rows in ws.iter_rows():
            for cell in rows:
                if cell.value is not None and "/" in str(cell.value):
                    cv = str(cell.value)
                    if "(1 п/гр)" in cv or "( 1 п/гр)" in cv or "(2 п/гр)" in cv or "(2 п/гр)" in cv:
                        if cv.count("/") == 1:
                            cell.fill = PatternFill(patternType='solid', fgColor="00CC00")
                if cell.value is not None and "/" in str(cell.value) or str(cell.value) == "Разговоры о важном":
                    clr_str = str(cell.value).replace("(1 п/гр)", "/").replace("(2 п/гр)", "/").replace("( 1 п/гр)", "/").replace("( 2 п/гр)", "/")
                    tech = clr_str.split("/")[-1] if clr_str != "Разговоры о важном" else "Куратор группы"
                    clr_str = clr_str.split("/")[0] if clr_str != "Разговоры о важном" else clr_str
                    
                    res = find_less(str(clr_str).strip())
                    res_cab = find_cabinet(str(cell.offset(0, 1).value))
                    res_tech = find_teacher(str(tech).strip())
                    if res == 0:
                        cell.fill = PatternFill(patternType='solid', fgColor='FFFFCC')
                    if res_cab == 0:
                        cell.offset(0, 1).fill = PatternFill(patternType='solid', fgColor='FF0000')
                    if res_tech == 0:
                        cell.fill = PatternFill(patternType='solid', fgColor='800080')
        print("БЕЖЕВЫЙ - ошибка в названии ПРЕДМЕТА")
        print("КРАСНЫЙ - ошибка в названии КАБИНЕТА")
        print("ФИОЛЕТОВЫЙ - ошибка в имени ПРЕПОДАВАТЕЛЯ")
        print("ЗЕЛЕНЫЙ - отсутствует \"/\"")
    except Exception as e:
        print(e)
    wb.save(filePath)


def main():
    # Получаем путь к файлу из аргумента командной строки
    if sys.argv is not None and len(sys.argv) > 1:
        filePath = sys.argv[1]
        check_teachers(filePath)

if __name__ == '__main__':
    try:
        main()
    except Exception as ex:
        print(ex)
        x = input()
x = input()