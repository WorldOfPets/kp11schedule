import openpyxl
import sys
from basedef import get_lesson
from openpyxl.styles import PatternFill

class Nakladka():
    def __init__(self, less_one, less_two) -> None:
        self.less_one = less_one
        self.less_two = less_two

def check_teachers(filePath):
    wb = openpyxl.load_workbook(filePath, read_only=False, keep_vba=True)

    ws = wb.active
    lessons = get_lesson(ws)
    sorted_lessons = []
    errors = []
    for less in lessons:
        #print(len(sorted_lessons))
        if len(sorted_lessons) == 0:
            sorted_lessons.append(less)
        elif len(sorted_lessons) == 1:
            if less == sorted_lessons[0]:
                nak = Nakladka(less, sorted_lessons[0])
                errors.append(nak)
            else:
                sorted_lessons.append(less)
        else:
            c = 0
            for l in sorted_lessons:
                if l == less:
                    c+=1
                    nak = Nakladka(less, l)
                    errors.append(nak)
            if c > 0:
                continue
            else:
                sorted_lessons.append(less)

    teach_err = []
    cab_err = []
    not_err = []
    for i in errors:
        if isinstance(i, Nakladka):
            if (i.less_one.teacher == i.less_two.teacher and i.less_one.cabinet == i.less_two.cabinet) or (str(i.less_one.cabinet) == "100" and str(i.less_two.cabinet == "100")):
                not_err.append(i)
            elif i.less_one.teacher == i.less_two.teacher and i.less_one.cabinet != i.less_two.cabinet:
                teach_err.append(i)
            elif i.less_one.cabinet == i.less_two.cabinet and i.less_one.teacher != i.less_two.teacher:
                cab_err.append(i)

    print(10 * "-" + "СКОРЕЕ ВСЕГО НЕ НАКЛАДКА" + 10 * "-")
    for i in not_err:
        print(f"Ячейки {i.less_one.group}/{i.less_two.group} Учителя {i.less_one.teacher} / {i.less_two.teacher} Кабинет {i.less_one.cabinet} / {i.less_two.cabinet}")
    print(10 * "-" + "НАКЛАДКА УЧИТЕЛЬ" + 10 * "-")
    for i in teach_err:
        print("ЖЕЛТЫЙ")
        ws[f"{i.less_one.coordinate}"].fill = PatternFill(patternType='solid', fgColor='FFFF00')
        ws[f"{i.less_two.coordinate}"].fill = PatternFill(patternType='solid', fgColor='FFFF00')
        print(f"Ячейки {i.less_one.coordinate}/{i.less_two.coordinate} Учителя {i.less_one.teacher} / {i.less_two.teacher} Кабинет {i.less_one.cabinet} / {i.less_two.cabinet}")
    print(10 * "-" + "НАКЛАДКА КАБИНЕТ" + 10 * "-")
    for i in cab_err:
        print("ОРАНЖЕВЫЙ")
        ws[f"{i.less_one.coordinate}"].fill = PatternFill(patternType='solid', fgColor='FF6600')
        ws[f"{i.less_two.coordinate}"].fill = PatternFill(patternType='solid', fgColor='FF6600')
        print(f"Ячейки {i.less_one.coordinate}/{i.less_two.coordinate} Учителя {i.less_one.teacher} / {i.less_two.teacher} Кабинет {i.less_one.cabinet} / {i.less_two.cabinet}")
    
    wb.save(str(filePath))
            
        


    

def main():
    # Получаем путь к файлу из аргумента командной строки
    if sys.argv is not None and len(sys.argv) > 1:
        filePath = sys.argv[1]
        check_teachers(filePath)

if __name__ == '__main__':
    try:
        main()
    except Exception as ex:
        print(ex)
        x = input()
x = input()