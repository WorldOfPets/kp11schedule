from typing import List
import openpyxl
from difflib import SequenceMatcher
import icalendar
import pytz
from dateutil.rrule import rrule, WEEKLY
from datetime import datetime, timedelta

wb = openpyxl.load_workbook("./excel/test2.xlsx")

ws = wb.active


days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
lessons = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]

class LessonKP11():
    teacher = ""
    cabinet = ""
    lesson_num = ""
    group = ""
    lesson_name = ""
    lesson_day = ""

    def __str__(self) -> str:
        return f"{self.teacher} {self.cabinet} {self.lesson_num} {self.group} {self.lesson_name} {self.lesson_day}"

    def check_teacher(self, out_teacher):
        matcher = SequenceMatcher(None, out_teacher, self.teacher)
        return True if matcher.ratio() > 0.80 else False
    
def check_teachers(teacher1, teacher2):
        matcher = SequenceMatcher(None, teacher1, teacher2)
        return True if matcher.ratio() > 0.80 else False
def get_group(row_id, col_id):
    for row in ws.iter_rows(min_row=row_id, max_row=row_id):
        for cell in row:
            if cell.value is not None and cell.column == col_id:
                return cell.value

def get_groups(row_id = 4):
    groups = []
    for row in ws.iter_rows(min_row=row_id, max_row=row_id):
        for cell in row:
            if cell.value is not None:
                groups.append(cell.value)
    return groups

def get_day(col_id, row_id):
    new_days = []
    for col in ws.iter_cols(min_col=col_id, max_col=col_id, max_row=row_id):
        for cell in col:
            if cell.value is not None and cell.value in days:
                #print(cell.value)
                new_days.append(cell.value)
    return new_days[-1] if new_days is not None else "понедельник"

def get_lesson_num(col_id, row_id) -> int:
    for col in ws.iter_cols(min_col=col_id, max_col=col_id, max_row=row_id, min_row=row_id-1):
        for cell in col:
            if cell.value is not None and cell.value in range(1, 11):
                return int(cell.value)

def get_teachers():
    teachers = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "/" in str(cell.value):
                print(len(teachers))
                if len(teachers) == 0:
                    teachers.append(cell.value.split("/")[-1].strip())
                elif len(teachers) == 1:
                    if not check_teachers(teachers[0], cell.value.split("/")[-1].strip()):
                        teachers.append(cell.value.split("/")[-1].strip())
                else:
                    c = 0
                    for i in teachers:
                        if check_teachers(i, cell.value.split("/")[-1].strip()):
                            c+=1
                        
                    if c > 0:
                        continue
                    else:
                        teachers.append(cell.value.split("/")[-1].strip())
                
    return set(teachers)

def get_cabinets():
    cabinets = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "/" in str(cell.value) or str(cell.value) == "Разговоры о важном":
                cabinets.append(cell.offset(0, 1).value)
    return set(cabinets)

def get_lesson() -> List[LessonKP11]:
    lessons_arr = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "/" in str(cell.value) or str(cell.value) == "Разговоры о важном":
                if cell.offset(0, 1).value is not None:
                    lesson = LessonKP11()
                    lesson.cabinet = cell.offset(0, 1).value
                    lesson.group = get_group(4, cell.column)
                    lesson.lesson_day = get_day(1, cell.row)
                    lesson.teacher = str(cell.value).split("/")[-1] if "/" in str(cell.value).strip() else "Куратор группы"
                    lesson.lesson_num = get_lesson_num(2, cell.row)
                    lesson.lesson_name = cell.value.split("/")[0] if "/" in str(cell.value) else cell.value
                    if "(1 п" in str(lesson.lesson_name) or "(2 п" in str(lesson.lesson_name):
                        lesson.lesson_name += "/гр)"
                    lessons_arr.append(lesson)
    return lessons_arr

def get_lesson_for_teacher(teacher, day) -> List[LessonKP11]:
    list_less_t = []
    for less in get_lesson():
        if less.check_teacher(teacher):
            list_less_t.append(less)
    
    print(len(list_less_t))
    return list_less_t

def get_sheet_for_teacher(teacher):
    sheet = []
    sheet.append(lessons[::-1])
    sheet_day = []
    lesss = get_lesson_for_teacher(teacher, 1)
    for d in days:
        sheet_day = []
        for less in lesss:
            if less.lesson_day == d:
                sheet_day.append(less.lesson_num)
        new_sheet_day = []
        for i in range(1, 11):
            if i in sheet_day:
                new_sheet_day.append(i)
            else:
                new_sheet_day.append("Свободен")          
        sheet.append(new_sheet_day[::-1])
    print(sheet)
    return sheet
            

def check_right_format():
    errors = []
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            if cell.value is not None and len(str(cell.value)) > 4:
                if ("(1 п" in cell.value or "(2 п" in cell.value) and cell.value != "Разговоры о важном" and str(cell.value).count("/") < 2:
                    errors.append("Ошибка отсутствия / в ячейки: " + cell.coordinate)
                if "." in cell.value and "  " in str(cell.value).strip():
                    #errors.append("Ошибка двойного пробела в ячейки: " + cell.coordinate)
                    continue
                if "." in cell.value and str(cell.value).count(".") == 1:
                    errors.append("Ошибка точки в инициалах перподавателя в ячейки: " + cell.coordinate)
                    
            elif cell.value is not None and len(str(cell.value)) < 3 and len(str(cell.value)) > 1:
                errors.append("Ошибка кабинета в ячейки: " + cell.coordinate)
    return errors


def get_lesson_for_group(group) -> List[LessonKP11]:
    list_less_g = []
    for less in get_lesson():
        if group in less.group:
            list_less_g.append(less)
    return list_less_g


def add_event(less_arr:List[LessonKP11], teacher, is_even):
    calendar = icalendar.Calendar()
    print(len(less_arr))
    for less in less_arr:
        this_wkst = days.index(less.lesson_day)

        dstart_e = datetime(2024, 9, 2 , 9, 0, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        dstart_e += timedelta(days=7) if is_even else timedelta(days=0)
        dstart_n = dstart_e + timedelta(minutes=55)
        
        if less.lesson_num % 2 == 1 and less.lesson_num <= 8:
            dstart_e += timedelta(hours=1*int(less.lesson_num)-1)
        elif less.lesson_num % 2 == 0 and less.lesson_num <= 8:
            dstart_n += timedelta(hours=1*int(less.lesson_num)-2)
        elif less.lesson_num == 9:
            dstart_e = datetime(2024, 9, 2 , 16, 50, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        elif less.lesson_num == 10:
            dstart_n = datetime(2024, 9, 2 , 17, 45, 0, tzinfo=pytz.timezone('Europe/Moscow'))


        # if less.lesson_num == 1:
        #     dstart = datetime(2024, 9, 2 , 9, 0, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 2:
        #     dstart = datetime(2024, 9, 2 , 9, 55, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 3:
        #     dstart = datetime(2024, 9, 2 , 11, 0, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 4:
        #     dstart = datetime(2024, 9, 2 , 11, 55, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 5:
        #     dstart = datetime(2024, 9, 2 , 13, 00, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 6:
        #     dstart = datetime(2024, 9, 2 , 13, 55, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 7:
        #     dstart = datetime(2024, 9, 2 , 15, 00, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 8:
        #     dstart = datetime(2024, 9, 2 , 15, 55, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 9:
        #     dstart = datetime(2024, 9, 2 , 16, 50, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        # if less.lesson_num == 10:
        #     dstart = datetime(2024, 9, 2 , 17, 45, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        dstart_e += timedelta(days=this_wkst)
        dstart_n += timedelta(days=this_wkst)
        dend = datetime(2024, 12, 31, 7, 0, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        listDate = list(rrule(freq=WEEKLY, 
                             interval=2, 
                             until=dend,
                             dtstart=dstart_e if less.lesson_num % 2 == 1 else dstart_n))

        for dateCurrent in listDate:
            event = icalendar.Event()
            event.add('summary', f"{less.group}/{less.cabinet}/{less.lesson_name}")
            event.add('location', f"{less.cabinet}")
            event.add('dtstart', dateCurrent)
            event.add('dtend', dateCurrent + timedelta(minutes=45))
            event.add('class', 'private')
            calendar.add_component(event)
    if is_even:
        with open(f'./chet/{teacher}_четная.ics', 'wb') as file:
            file.write(calendar.to_ical())
    else:
        with open(f'./nech/{teacher}_нечетная.ics', 'wb') as file:
            file.write(calendar.to_ical())

def add_event_group(less_arr:List[LessonKP11], group_name, is_even):
    calendar = icalendar.Calendar()
    print(len(less_arr))
    for less in less_arr:
        this_wkst = days.index(less.lesson_day)

        dstart_e = datetime(2024, 9, 2 , 9, 0, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        dstart_e += timedelta(days=7) if is_even else timedelta(days=0)
        dstart_n = dstart_e + timedelta(minutes=55)
        
        if less.lesson_num % 2 == 1 and less.lesson_num <= 8:
            dstart_e += timedelta(hours=1*int(less.lesson_num)-1)
        elif less.lesson_num % 2 == 0 and less.lesson_num <= 8:
            dstart_n += timedelta(hours=1*int(less.lesson_num)-2)
        elif less.lesson_num == 9:
            dstart_e = datetime(2024, 9, 2 , 16, 50, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        elif less.lesson_num == 10:
            dstart_n = datetime(2024, 9, 2 , 17, 45, 0, tzinfo=pytz.timezone('Europe/Moscow'))

        dstart_e += timedelta(days=this_wkst)
        dstart_n += timedelta(days=this_wkst)
        dend = datetime(2024, 12, 31, 7, 0, 0, tzinfo=pytz.timezone('Europe/Moscow'))
        listDate = list(rrule(freq=WEEKLY, 
                             interval=2, 
                             until=dend,
                             dtstart=dstart_e if less.lesson_num % 2 == 1 else dstart_n))

        for dateCurrent in listDate:
            event = icalendar.Event()
            event.add('summary', f"{less.cabinet}/{less.teacher}/{less.lesson_name}")
            event.add('location', f"{less.cabinet}")
            event.add('dtstart', dateCurrent)
            event.add('dtend', dateCurrent + timedelta(minutes=45))
            event.add('class', 'public')
            calendar.add_component(event)
    if is_even:
        with open(f'./chet/{group_name}_четная.ics', 'wb') as file:
            file.write(calendar.to_ical())
    else:
        with open(f'./nech/{group_name}_нечетная.ics', 'wb') as file:
            file.write(calendar.to_ical())