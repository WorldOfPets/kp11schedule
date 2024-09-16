import openpyxl
import sys
from basedef import get_lesson, get_day, get_lesson_num
from dbfunc import get_cab_db
from openpyxl.styles import PatternFill

days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
num = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
class ScheObj():
    row_id = 0
    col_id = 0
    is_tec = False

def find_cabinet(filePath, row_id, col_id):
    wb = openpyxl.load_workbook(filePath, read_only=False, keep_vba=True)

    ws = wb.active
    cv = str(ws.cell(row=int(row_id), column=int(col_id)).value).strip()
    teacher = str(cv.split("/")[-1]).strip()
    teacher = teacher.split(" ")[0]
    print(teacher)
    she_arr = []
    lesson = get_lesson(ws)
    for less in lesson:
        if teacher in less.teacher:
            scheObj = ScheObj()
            scheObj.row_id = less.lesson_num + 1
            scheObj.col_id = days.index(less.lesson_day) + 2
            scheObj.is_tec = True
            print(scheObj.row_id)
            print(scheObj.col_id)
            print(scheObj.is_tec)
            she_arr.append(scheObj)


    ws = wb.create_sheet(teacher)
    ws = wb[teacher]
    ws.cell(1, 1).value = teacher
    for i in range(2, len(days)+2):
        ws.cell(1, i).value = days[i-2]
    for i in range(2, 12):
        ws.cell(i, 1).value = i-1

    for obj in she_arr:
        ws.cell(obj.row_id, obj.col_id).value = "Занят"
        ws.cell(obj.row_id, obj.col_id).fill = PatternFill(patternType='solid', fgColor='FF3300')
    for rows in ws.iter_rows(max_row=11, max_col=7):
        for cell in rows:
            if cell.value is None:
                cell.value = "Свободен"
                cell.fill = PatternFill(patternType='solid', fgColor='00FF00')
                  

   
    
    wb.save(filePath)


def main():
    if sys.argv is not None and len(sys.argv) > 1:
        filePath = sys.argv[1]
        find_cabinet(filePath, sys.argv[2], sys.argv[3])

if __name__ == '__main__':
    main()
x = input()